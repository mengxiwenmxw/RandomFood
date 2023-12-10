import tkinter as tk
import json
import random
import os
from tkinter import filedialog
import openpyxl


def save_food(filename, data):
    with open(filename, "w") as json_file:
        json.dump(data, json_file, indent=4)


def load_all_food(filename):
    with open(filename, 'r') as file:
        data = json.load(file)
        # print(type(data))
    return data


def open_file():
    suffix=['.xlsx','.xlsm','.xltx','.xltm','.xls','.xlt']
    data=[]
    filepath = filedialog.askopenfilename()
    extension = os.path.splitext(filepath)[1].lower()
    if extension in suffix:
        wb = openpyxl.load_workbook(filename=filepath)
        # 获取工作表名
        sheet_name = wb.sheetnames[0]
        # 选择工作表
        sheet = wb[sheet_name]
        for rows in sheet.iter_rows(min_row=0,values_only=True):
            for row in rows:
                if row:
                    data.append(row)
    return data


class Gui:
    def __init__(self, master):
        self.output_box = None
        self.menu_window = None
        self.show_food = None
        self.root = master
        self.root.title('RandomFood')
        self.root.iconbitmap('photo.ico')
        self.root.geometry("400x300")
        self.root.resizable(False, False)

        # food
        self.all_food_path = 'food_file.json'
        self.check_food_path = 'check_food.json'
        self.all_food = load_all_food(self.all_food_path)
        self.check_food = load_all_food(self.check_food_path)
        self.check_boxs = []
        self.all_food_change = {v: k for k, v in self.all_food.items()}
        self.swapped_check_food = {v: k for k, v in self.check_food.items()}
        # print(self.all_food_change)

        self.flag_stop = False

    def import_food(self):
        self.all_food = load_all_food(self.all_food_path)
        self.check_food = load_all_food(self.check_food_path)
        self.check_boxs = []
        self.all_food_change = {v: k for k, v in self.all_food.items()}
        self.swapped_check_food = {v: k for k, v in self.check_food.items()}

    def set_window(self):
        main_menu = tk.Menu(self.root, tearoff=False)
        self.root.config(menu=main_menu)

        # select button
        sel_btn_start = tk.Button(self.root, command=self.btn_callback, text='start', cursor='hand2',
                                  activebackground='blue',
                                  activeforeground='white')
        sel_btn_start.pack(side='left')
        sel_btn_stop = tk.Button(self.root, command=self.btn_callback_stop, text='stop', cursor='hand2',
                                 activebackground='blue',
                                 activeforeground='white')
        sel_btn_stop.pack(side='right')
        # food output
        frame_output = tk.LabelFrame(self.root, text='今天吃什么', padx=70, pady=90)
        frame_output.pack(fill='none', expand=True)
        self.output_box = tk.Text(frame_output, wrap=tk.WORD, state='normal', bg='#667766', fg='white', undo=True)
        self.output_box.pack(fill='none', pady=10, padx=10, expand=True)
        self.output_box.insert(tk.END, f"output window")
        # menubar
        add_food = tk.Menu(main_menu, tearoff=False)
        select_food = tk.Menu(main_menu, tearoff=False)
        """
        set food:add,remove,show
        """
        main_menu.add_cascade(label='food lib', menu=add_food)
        add_food.add_command(label='edit food lib', command=self.show_edit_food)
        add_food.add_command(label='import food from excel', command=self.get_file_food)
        """
        select food:checkbutton
        """
        main_menu.add_cascade(label='Select food that you wanna eat', menu=select_food)
        select_food.add_command(label='menu', command=self.show_menu)

    def get_file_food(self):
        datas = open_file()
        if datas:
            index = 1
            len1=len(self.all_food)
            len2=len(self.check_food)
            for data in datas:
                flag_re = False
                for key in self.all_food.keys():
                    if str(data) == key:
                        flag_re = True
                if not flag_re:
                    self.all_food[str(data)] = str(len1 + index)
                    self.check_food[str(data)] = str(len2 + index)
                    index += 1
            save_food(filename=self.all_food_path,data=self.all_food)
            save_food(filename=self.check_food_path,data=self.check_food)
            self.import_food()


    def show_menu(self):
        global listbox, items_vars, canvas
        self.menu_window = tk.Toplevel(self.root)
        self.menu_window.title('menu')
        self.menu_window.iconbitmap('photo.ico')
        #
        ok_btn = tk.Button(self.menu_window, command=self.check_finish, text='ok', cursor='hand2',
                           activebackground='blue',
                           activeforeground='white')
        ok_btn.pack(side=tk.LEFT, anchor=tk.SW)
        self.menu_window.bind('<Return>', self.check_finish)
        scrollbar = tk.Scrollbar(self.menu_window)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)
        scrollbar.bind('<Return>', self.check_finish)
        listbox = tk.Listbox(self.menu_window, yscrollcommand=scrollbar.set)
        if self.all_food != ():
            for item in self.all_food.keys():
                listbox.insert(tk.END, item)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH)
        listbox.bind('<Return>', self.check_finish)
        scrollbar.config(command=listbox.yview)

        scrollbar_check = tk.Scrollbar(self.menu_window)
        scrollbar_check.pack(side=tk.RIGHT, fill=tk.Y)
        canvas = tk.Canvas(self.menu_window, yscrollcommand=scrollbar_check.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_check.config(command=canvas.yview)
        canvas.bind("<MouseWheel>", self.on_mousewheel)
        frame = tk.Frame(canvas)
        canvas.create_window((30, 0), anchor=tk.NW, window=frame)

        items_vars = {}

        if self.all_food != ():
            for item in self.all_food.keys():
                var = tk.BooleanVar()
                items_vars[item] = var
                self.check_boxs.append(var)
                checkbutton = tk.Checkbutton(frame, text=item, variable=var)
                if any(item == d for d, k in self.check_food.items()):
                    checkbutton.select()
                checkbutton.pack()
            listbox.bind("<<ListboxSelect>>", self.on_select)

        # 绑定 Scrollbar 和 Canvas
        canvas.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

    def on_mousewheel(self, event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def on_select(self, event):
        selected_items = listbox.get(listbox.curselection())
        if items_vars[selected_items].get() == 1:
            items_vars[selected_items].set(False)
        else:
            items_vars[selected_items].set(True)

    def check_finish(self, *args):
        # print('check ok')
        checked = []
        i = 1
        for var in self.check_boxs:
            if var.get():
                checked.append(self.all_food_change[str(i)])
            i += 1
        # print(checked)
        self.check_boxs = []
        self.check_food = {}
        i = 1
        for check in checked:
            self.check_food[check] = str(i)
            i += 1
        # print(self.check_food)
        save_food(self.check_food_path, self.check_food)
        self.import_food()
        self.menu_window.destroy()

    def show_edit_food(self):
        global list_
        self.show_food = tk.Toplevel(self.root)
        self.show_food.title('food')
        self.show_food.geometry('200x300')
        self.show_food.iconbitmap('photo.ico')
        self.show_food.resizable(False, False)
        # Scrollbar
        scr = tk.Scrollbar(self.show_food)
        scr.pack(side='right', fill='y')
        # show
        list_ = tk.Listbox(self.show_food, yscrollcommand=scr.set, selectmode='extended')
        list_.pack()
        list_.bind('<Delete>', self.remove_food)
        if self.all_food != ():
            for food in self.all_food.keys():
                list_.insert('end', food)
        else:
            # print('null')
            list_.insert('null')
        scr.config(command=list_.yview)
        # edit
        btn_frame = tk.Frame(self.show_food)
        btn_frame.pack(fill=tk.X, padx=10, pady=10, expand=True)
        btn_add = tk.Button(btn_frame, command=self.new_food, text='add food', cursor='hand2',
                            activebackground='blue',
                            activeforeground='white')
        btn_add.grid(row=0, column=0, sticky='ne')
        btn_remove = tk.Button(btn_frame, command=self.remove_food, text='remove food', cursor='hand2',
                               activebackground='blue',
                               activeforeground='white')
        btn_remove.grid(row=0, column=2, sticky='e')

    def btn_callback(self):
        if not self.flag_stop:
            self.flag_stop = True
            self.root.after(200, self.random_show)  # 每隔2秒调用一次 update_text()

    def btn_callback_stop(self):
        self.flag_stop = False

    def random_show(self):
        if self.flag_stop:
            random_integer = random.randint(1, len(self.check_food))
            # print(random_integer)
            # random_text = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(10))
            self.output_box.delete(1.0, tk.END)  # 清空 Text 控件中的内容
            self.output_box.insert(tk.END, self.swapped_check_food[str(random_integer)] + "\n")
            self.output_box.tag_configure("center", justify="center", font=("Arial", 15))  # 设置一个居中对齐的标签配置
            self.output_box.tag_add("center", "1.0", "end")  # 应用标签配置到整个文本范围
            self.root.after(20, self.random_show)

    def new_food(self):
        global new_food_name, mes
        new_food_win = tk.Toplevel(self.show_food)
        new_food_win.title('add food')
        new_food_win.geometry('200x150')
        new_food_win.iconbitmap('photo.ico')
        new_food_win.resizable(False, False)
        # entry
        mes = tk.Label(new_food_win, text='*food name:')
        mes.pack()
        new_food_name = tk.Entry(new_food_win)
        new_food_name.pack()
        new_food_name.bind("<Return>", self.save_new_food)
        finish_btn = tk.Button(new_food_win, command=self.add_finish, text='finished', cursor='hand2',
                               activebackground='blue',
                               activeforeground='white')
        finish_btn.pack(side='bottom')
        save_btn = tk.Button(new_food_win, command=self.save_new_food, text='save food', cursor='hand2',
                             activebackground='blue',
                             activeforeground='white')
        save_btn.pack(side='bottom')

        # print('add ok')

    def add_finish(self):
        self.show_food.destroy()
        self.show_edit_food()

    def save_new_food(self, *args):
        if new_food_name.get() != '':
            flag_re = False
            for key in self.all_food.keys():
                if new_food_name.get() == key:
                    flag_re = True
                    mes.config(text=f'repeated food')
            if not flag_re:
                self.all_food[new_food_name.get()] = str(len(self.all_food) + 1)
                self.check_food[new_food_name.get()] = str(len(self.check_food) + 1)
                # print(self.all_food)
                save_food(filename=self.all_food_path, data=self.all_food)
                save_food(filename=self.check_food_path, data=self.check_food)
                mes.config(text=f'successfully add {new_food_name.get()}:')
            new_food_name.delete(0, tk.END)

        else:
            mes.config(text=f'NULL,entry again:')
        self.show_food.update()
        self.import_food()

    def remove_food(self, *args):
        # print(list_.curselection())
        del_keys = []
        if list_.curselection() != ():
            for num in list_.curselection()[::-1]:
                list_.delete(num)
                for key in self.all_food.keys():
                    if self.all_food[key] == str(num + 1):
                        del_keys.append(key)
                        break
            for del_key in del_keys:
                del self.all_food[del_key]
                if del_key in self.check_food.keys():
                    del self.check_food[del_key]
            i, j = 1, 1
            for k, v in self.check_food.items():
                self.check_food[k] = str(i)
                i += 1
            # print(num_min)
            for key in self.all_food.keys():
                self.all_food[key] = str(j)
                j += 1
            save_food(self.all_food_path, self.all_food)
            save_food(self.check_food_path, self.check_food)
            self.import_food()


if __name__ == '__main__':
    root = tk.Tk()
    window = Gui(root)
    window.set_window()
    # window
    root.mainloop()
